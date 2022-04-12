import os
import json as jsonlib
from abc import ABC, abstractmethod
from typing import Any, Optional
import openpyxl
from openpyxl.drawing.image import Image
from ..utils import workspace


class Export(ABC):
    def __init__(self, output: str):
        """

        Args:
            output (str): Output File Path
        """
        self.file_ext_check(output)
        self.data = self.load(output)
        self.output = output

    def list_exist_dirs(self):
        perk_dir_list = []
        for root, dirs, files in os.walk(os.path.join(workspace(), "data", "wish_list")):
            if "icons" in dirs and "data.json" in files:
                perk_dir_list.append(root)
        return perk_dir_list

    def get_perk_icon(self, perk_id: str) -> str:
        perk_icon_dir = os.path.join(workspace(), "data", "perks", "icons")
        perk_icon_file = os.path.join(perk_icon_dir, f"{perk_id}.jpg")
        if os.path.isfile(perk_icon_file) is False:
           perk_icon_file = None
        return perk_icon_file

    @abstractmethod
    def file_ext_check(self, output: str):
        pass

    @abstractmethod
    def load(self, output: str) -> Any:
        pass

    @abstractmethod
    def process(self):
        pass

    @abstractmethod
    def save(self):
        pass


class ExcelExport(Export):
    def file_ext_check(self, output: str):
        _ext = os.path.splitext(output)[-1]
        if _ext not in [".xlsx", ".xls"]:
            raise TypeError(f"Output File Ext Invalid, {_ext}\n"
                            f"Expect Ext is one of [xlsx|xls]")

    def load(self, output: str) -> openpyxl.Workbook:
        if os.path.isfile(output):
            wb = openpyxl.load_workbook(output)
        else:
            wb = openpyxl.Workbook()
        return wb

    def process(self):
        ws = self.data.active
        # ws = openpyxl.Workbook().active
        index = 1
        for element in self.list_exist_dirs():
            icon = os.path.join(element, "icons", "icon.jpg")
            watermark = os.path.join(element, "icons", "watermark.jpg")
            data_file = os.path.join(element, "data.json")
            with open(data_file, "r") as fp:
                json_data = jsonlib.load(fp)

            wish_list = json_data['perks']
            for index_1, wish_perk in enumerate(wish_list):
                _index = index + index_1
                row = [
                    '',  # water marks in A
                    '',  # icon image in B
                    json_data['category'],
                    json_data['weapon_type'],
                    json_data['damage_type'],
                    json_data['name'],
                ]
                ws.append(row)
                watermark_img = Image(watermark)
                icon_img = Image(icon)
                ws.add_image(watermark_img, f"A{_index}")
                ws.add_image(icon_img, f"B{_index}")

                current_row = ws.rows[_index]
                for perk in wish_perk['perks']:
                    perk_icon = self.get_perk_icon(perk['hash'])
                    if perk_icon is None:
                        current_row
                    perk_img = Image(perk_icon)

    def save(self):
        pass


